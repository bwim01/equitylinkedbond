### STEP0. 라이브러리 가져오기
import requests
import zipfile
import os
import pandas as pd
import numpy as np
import re
import math
import pickle
import warnings
import streamlit as st
from datetime import datetime, timedelta, date
import json
from streamlit_lottie import st_lottie

from bs4 import BeautifulSoup
from io import BytesIO
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, PatternFill, colors, Color, Font

warnings.filterwarnings(action='ignore')
API_KEY = '71b753d5fc4afd62bcafda3e1ae36fc53aad1cff'

### STEP1. 주요사항보고서 중 전환사채권, 신주인수권부사채권, 교환사채권 정보 가져오기
def get_rcept_no(bgn_de, end_de) :
    rcept_info = []
    url = 'https://opendart.fss.or.kr/api/list.xml'
    params = {'crtfc_key': API_KEY
              , 'bgn_de' : bgn_de
              , 'end_de' : end_de
              , 'sort_mth' : 'asc' # 정렬방법(오름차순)
              , 'pblntf_detail_ty': 'B001' # 주요사항보고서
              , 'last_reprt_at':'N'} # 최종보고서 검색여부(아니요)
    response = requests.get(url, params=params, verify=False)
    soup = BeautifulSoup(response.content, features='xml')
    total_page = soup.find('total_page').get_text()
    
    for i in range(1, int(total_page) + 1):
        params = {'crtfc_key': API_KEY
                  , 'bgn_de': bgn_de
                  , 'end_de': end_de
                  , 'sort_mth' : 'asc' # 정렬방법(오름차순)
                  , 'pblntf_detail_ty': 'B001' # 주요사항보고서
                  , 'page_no': str(i) 
                  , 'last_reprt_at':'N'} # 최종보고서 검색여부(아니요)
        response = requests.get(url, params=params, verify=False)
        soup = BeautifulSoup(response.content, features='xml')
        rcept_names = ['주요사항보고서(전환사채권발행결정)', '주요사항보고서(신주인수권부사채권발행결정)', '주요사항보고서(교환사채권발행결정)']
        for c in soup.find_all('list'):
            if c.report_nm.get_text() in rcept_names:
                rcept_info.append(c.rcept_no.get_text()+'_A'+c.stock_code.get_text()+'_'+c.corp_cls.get_text())
        
    print('보고서수 : ', len(rcept_info))
    st.write('보고서 수 : ', len(rcept_info)) 
    return(rcept_info)

### STEP2. 신고서 데이터 수집하기
def get_corp_docu(rcept_no):
    url = 'https://opendart.fss.or.kr/api/document.xml'
    params = {'crtfc_key': API_KEY, 'rcept_no': rcept_no}
    response = requests.get(url, params=params)
    row = {}
    
    try:
        zf = zipfile.ZipFile(BytesIO(response.content))
        z_list = zf.namelist()
        file = zf.read(z_list[0]) 

        soup = BeautifulSoup(file, 'html.parser', from_encoding='utf-8')
        company_nm = soup.find('company-name').get_text() 
        doc_nm = '전환사채권' if '전환사채권' in soup.find('document-name').get_text() else ('신주인수권부사채권' if '신주인수권' in soup.find('document-name').get_text() else '교환사채권')
        table = soup.find('table-group', attrs={'aclass':'CB_PUB'}) if doc_nm=='전환사채권' else (soup.find('table-group', attrs={'aclass':'BW_PUB'}) if doc_nm=='신주인수권부사채권' else soup.find('table-group', attrs={'aclass':'EB_PUB'}))
        
        rcept_dt = rcept_no[:8] 
        pym_dt = table.find('tu', attrs={'aunit':'PYM_DT'}).get('aunitvalue')         
        dnm_sum = table.find('te', attrs={'acode':'DNM_SUM'}).get_text()  
        prft_rate = table.find('te', attrs={'acode':'PRFT_RATE'}).get_text() 
        lst_rtn_rt = table.find('te', attrs={'acode':'LST_RTN_RT'}).get_text() 
        exp_dt = table.find('tu', attrs={'aunit':'EXP_DT'}).get('aunitvalue') 
        sb_bgn_dt = table.find('tu', attrs={'aunit':'SB_BGN_DT'}).get('aunitvalue') 
        try:
            min_prc = table.find('te', attrs={'acode':'MIN_PRC'}).get_text()
        except:
            min_prc = ''
        exe_prc = table.find('te', attrs={'acode':'EXE_PRC'}).get_text() 
        rpt_sm_yn = table.find('tu', attrs={'aunit':'RPT_SM_YN'}).get_text()
        
        table1 = soup.find('table-group', attrs={'aclass':'CRP_ISSU'})
        num = len(table1.find_all('te', attrs={'acode':'ISSU_NM'}))
        issu_nm = []
        for i in range(num):
            issu_nm.append(table1.find_all('te', attrs={'acode':'ISSU_NM'})[i].get_text())
                
        row = {'발행사':company_nm, '구분': doc_nm, '공시일':rcept_dt, '납입일':pym_dt, '전자등록총액':dnm_sum, '표면이자율(%)':prft_rate,
                '만기이자율(%)':lst_rtn_rt, '사채만기일':exp_dt, '전환청구시작일':sb_bgn_dt, '최저조정가액':min_prc, '전환가액':exe_prc ,
                '제출대상여부':rpt_sm_yn, '발행대상자명':issu_nm}
    except Exception as e:
        print(rcept_no+'_Error!_'+str(e))        
    return row

### STEP3. 원하는 형식으로 정리하기
def get_report(info):
    rows=[]
    for i in reversed(range(len(info))):
        try:
            data = get_corp_docu(info[i][:14])
            rcept_dt = datetime.strptime(data['공시일'],'%Y%m%d')
            pym_dt = datetime.strptime(data['납입일'],'%Y%m%d')
            company_cd = '' if info[i][15:22] == 'A_E' else info[i][15:22]
            company_nm = data['발행사'].replace('(주)','').replace('㈜','').replace(' 주식회사','').replace('주식회사 ','').replace('주식회사','').strip()
            company_type = '유가증권' if info[i][-1]=='Y' else ('코스닥' if info[i][-1]=='K' else ('코넥스' if info[i][-1]=='N' else '기타법인' ))  
            type1 = '공모' if data['제출대상여부'] =='예' else '사모'
            type2 = '영구' if round((datetime.strptime(data['사채만기일'],'%Y%m%d')-datetime.strptime(data['납입일'],'%Y%m%d')).days/365,1)>=30 else ''
            type3 = 'CB' if data['구분'] == '전환사채권' else ('BW' if data['구분'] == '신주인수권부사채권' else 'EB')
            dnm_sum = int(data['전자등록총액'].replace('\n','').replace(',',''))/1000000
            prft_rate = '-' if data['표면이자율(%)'] == '-' else float(data['표면이자율(%)'])/100
            lst_rtn_rt = '-' if data['만기이자율(%)'] == '-' else float(data['만기이자율(%)'])/100
            exp_dt = datetime.strptime(data['사채만기일'],'%Y%m%d')
            exp_year = (str(round((datetime.strptime(data['사채만기일'],'%Y%m%d')-datetime.strptime(data['납입일'],'%Y%m%d')).days/365,1))+'년').replace('.0','')
            sb_bgn_dt = datetime.strptime(data['전환청구시작일'],'%Y%m%d')
            value = ['', '-']
            prc_ox = '' if data['최저조정가액'] in value else round((int(data['최저조정가액'].replace(',',''))/int(data['전환가액'].replace(',','')))*100)
            if prc_ox == '':
                prc_o = ''
                prc_x = ''
            elif prc_ox >= 70:       
                prc_o = ''
                prc_x = float(data['최저조정가액'].replace(',',''))/float(data['전환가액'].replace(',',''))
            else :
                prc_o = float(data['최저조정가액'].replace(',',''))/float(data['전환가액'].replace(',',''))
                prc_x = ''
            rpt_sm_yn = data['제출대상여부'].replace('아니오','면제').replace('예','해당')
            issu_nm=[]
            for j in range(len(data['발행대상자명'])):
                if not re.search(r'\(.*신탁.*\)', data['발행대상자명'][j]) :
                    issu_nm.append(data['발행대상자명'][j].replace('\n','').replace('(주)','').replace('㈜','').replace(' 주식회사','').replace('주식회사 ','').replace('주식회사','').strip())
            issu_nms = ', '.join(issu_nm)
            rept_no = info[i][:14]
            
            row = {'공시일':rcept_dt, '납입일':pym_dt, '발행사(Code)':company_cd, '발행사':company_nm, '법인유형':company_type, '종류':type1+type2+type3,
                  '전자등록총액(백만원)':dnm_sum, 'Coupon':prft_rate, 'YTM':lst_rtn_rt, 'YTP':'', 'Put 시작':'', 'Put 주기':'', 'YTC':'', 
                  'Call 시작':'','Call 종료':'','Call 주기':'', 'Call %':'', '만기일':exp_dt, '만기':exp_year, '전환개시':sb_bgn_dt, '할증율':'',
                  '하향Refixing(액면가x)':prc_x, '하향Refixing(액면가o)':prc_o,'상향Refixing':'','Refixing시작':'','신고서':rpt_sm_yn, '투자자':issu_nms, '보고서':rept_no,}
            rows.append(row)
            print(info[i])
        except Exception as e:
                print(info[i]+'_Error!_'+str(e))
                st.write('<p style="font-size:14px; color:red">'+'- 문서 '+info[i][:14]+'에서 오류 발생! 데이터솔루션부에 문의하세요.</p>',unsafe_allow_html=True)
    result = pd.DataFrame(rows)
    return(result)

### STEP4. 웹페이지 레이아웃 및 엑셀 형식 설정하기
# 애니메이션 및 보고서 제목 삽입
def load_lottie():
    with open('./resources/report.json', 'r', encoding='utf-8-sig') as st_json:
        return json.load(st_json)

empty1, col1, col2 = st.columns([0.05, 0.3, 0.8])
with empty1:
    st.empty()
with col1:
    lottie = load_lottie()
    st_lottie(lottie, speed=1, loop=True, width=150, height=150, )
with col2:
    ''
    ''
    st.title('주식연계채권 발행현황')

# 날짜 선택
start_date = st.date_input('시작일', value=date.today(), max_value = date.today())
max_date = min(start_date+timedelta(days=31), date.today())
end_date = st.date_input('종료일', value=start_date, min_value = start_date, max_value = max_date)

# 조회 및 다운 버튼 생성
if st.button("조회"):
    bgn_de = datetime.strftime(start_date,'%Y%m%d')
    end_de = datetime.strftime(end_date,'%Y%m%d')
    info = get_rcept_no(bgn_de, end_de) 
    result = get_report(info)

    wb = Workbook()
    ws = wb.active
    for r in dataframe_to_rows(result, index=False, header=True):
        ws.append(r)
    for column_cells in ws.columns:
        for cell in ws[column_cells[0].column_letter]:
            cell.font = Font(size=9)
            if column_cells[0].column_letter in ['A','B','R','T'] :
                cell.number_format = 'YYYY-MM-DD'
            elif column_cells[0].column_letter in ['H','I']:
                cell.number_format = '0.00%'
            elif column_cells[0].column_letter in ['V','W']:
                cell.number_format = '0%'
            elif column_cells[0].column_letter == 'G':
                cell.number_format = '#,##0'   
    wb.save('주식연계채권_'+bgn_de+'_'+end_de+'.xlsx')
    st.dataframe(result)
    with open('주식연계채권_'+bgn_de+'_'+end_de+'.xlsx', 'rb') as f:
            data = f.read()
            st.download_button(label='다운', data=data, file_name='주식연계채권_'+bgn_de+'_'+end_de+'.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
